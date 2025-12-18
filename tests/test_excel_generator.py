"""
BudgetGuard ZAR - Excel Generator Tests.

Unit tests for ExcelReporter class.
Tests ensure correct sheet creation, data population,
and formatting application.
"""

import tempfile
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.excel_generator import ExcelReporter
from src.schema import (
    AnalysisSnapshot,
    Campaign,
    CampaignAnalysis,
    RiskLevel,
)


def create_test_snapshot(num_campaigns: int = 3) -> AnalysisSnapshot:
    """Creates a test snapshot with specified number of campaigns."""
    campaigns = []
    risk_levels = [RiskLevel.HEALTHY, RiskLevel.WARNING, RiskLevel.CRITICAL]
    
    for i in range(num_campaigns):
        budget = Decimal(str(10000 * (i + 1)))
        spend = Decimal(str(3000 * (i + 1)))
        
        campaign = Campaign(
            name=f"Campaign_{i + 1}",
            monthly_budget=budget,
            current_spend=spend
        )
        
        analysis = CampaignAnalysis(
            campaign=campaign,
            rds=Decimal(str(500 * (i + 1))),
            spend_percentage=Decimal("30.00"),
            time_percentage=Decimal("58.06"),
            risk_level=risk_levels[i % len(risk_levels)],
            days_remaining=14
        )
        campaigns.append(analysis)
    
    total_budget = sum(ca.campaign.monthly_budget for ca in campaigns)
    total_spend = sum(ca.campaign.current_spend for ca in campaigns)
    critical_count = sum(1 for ca in campaigns if ca.risk_level == RiskLevel.CRITICAL)
    warning_count = sum(1 for ca in campaigns if ca.risk_level == RiskLevel.WARNING)
    
    return AnalysisSnapshot(
        timestamp=datetime(2024, 12, 18, 14, 30, 0),
        version="0.1.0",
        campaigns=campaigns,
        total_budget=total_budget,
        total_spend=total_spend,
        critical_count=critical_count,
        warning_count=warning_count
    )


class TestExcelReporterUnit:
    """Unit tests for ExcelReporter."""

    def setup_method(self) -> None:
        """Initialise ExcelReporter for each test."""
        self.reporter = ExcelReporter()

    def test_generate_report_creates_file(self) -> None:
        """Verify report generation creates a file."""
        snapshot = create_test_snapshot()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            
            self.reporter.generate_report(snapshot, output_path)
            
            assert output_path.exists()

    def test_report_has_two_sheets(self) -> None:
        """Verify report has Finance Summary and Campaign Deep-Dive sheets."""
        snapshot = create_test_snapshot()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            self.reporter.generate_report(snapshot, output_path)
            
            workbook = load_workbook(output_path)
            sheet_names = workbook.sheetnames
            
            assert "Finance Summary" in sheet_names
            assert "Campaign Deep-Dive" in sheet_names
            assert len(sheet_names) == 2

    def test_summary_sheet_has_title(self) -> None:
        """Verify Finance Summary sheet has title."""
        snapshot = create_test_snapshot()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            self.reporter.generate_report(snapshot, output_path)
            
            workbook = load_workbook(output_path)
            ws = workbook["Finance Summary"]
            
            assert "BudgetGuard ZAR" in str(ws["A1"].value)

    def test_summary_sheet_has_portfolio_metrics(self) -> None:
        """Verify Finance Summary has portfolio metrics."""
        snapshot = create_test_snapshot()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            self.reporter.generate_report(snapshot, output_path)
            
            workbook = load_workbook(output_path)
            ws = workbook["Finance Summary"]
            
            # Check for key metric labels
            values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
            
            assert any("Total Portfolio Budget" in v for v in values)
            assert any("Total Spend" in v for v in values)
            assert any("RDS" in v for v in values)

    def test_summary_sheet_has_risk_counts(self) -> None:
        """Verify Finance Summary has risk count table."""
        snapshot = create_test_snapshot()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            self.reporter.generate_report(snapshot, output_path)
            
            workbook = load_workbook(output_path)
            ws = workbook["Finance Summary"]
            
            values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
            
            assert any("CRITICAL" in v for v in values)
            assert any("WARNING" in v for v in values)
            assert any("HEALTHY" in v for v in values)

    def test_detail_sheet_has_headers(self) -> None:
        """Verify Campaign Deep-Dive has correct headers."""
        snapshot = create_test_snapshot()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            self.reporter.generate_report(snapshot, output_path)
            
            workbook = load_workbook(output_path)
            ws = workbook["Campaign Deep-Dive"]
            
            headers = [cell.value for cell in ws[1]]
            
            assert "Campaign" in headers
            assert "Monthly Budget" in headers
            assert "Current Spend" in headers
            assert "RDS" in headers
            assert "Risk Status" in headers

    def test_detail_sheet_row_count_matches_campaigns(self) -> None:
        """Verify Campaign Deep-Dive has correct number of data rows."""
        num_campaigns = 5
        snapshot = create_test_snapshot(num_campaigns)
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            self.reporter.generate_report(snapshot, output_path)
            
            workbook = load_workbook(output_path)
            ws = workbook["Campaign Deep-Dive"]
            
            # Count non-empty rows (excluding header)
            data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value)
            
            assert data_rows == num_campaigns

    def test_detail_sheet_campaign_names_present(self) -> None:
        """Verify all campaign names appear in detail sheet."""
        snapshot = create_test_snapshot(3)
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            self.reporter.generate_report(snapshot, output_path)
            
            workbook = load_workbook(output_path)
            ws = workbook["Campaign Deep-Dive"]
            
            campaign_names = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]
            
            assert "Campaign_1" in campaign_names
            assert "Campaign_2" in campaign_names
            assert "Campaign_3" in campaign_names

    def test_currency_formatting_applied(self) -> None:
        """Verify ZAR currency formatting is applied."""
        snapshot = create_test_snapshot(1)
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            self.reporter.generate_report(snapshot, output_path)
            
            workbook = load_workbook(output_path)
            ws = workbook["Campaign Deep-Dive"]
            
            # Check budget column (column B) formatting
            budget_cell = ws.cell(row=2, column=2)
            
            assert "R" in budget_cell.number_format or "#,##0" in budget_cell.number_format

    def test_generate_filename(self) -> None:
        """Verify filename generation format."""
        filename = self.reporter.generate_filename("budget_report")
        
        assert filename.startswith("budget_report_")
        assert filename.endswith(".xlsx")

    def test_report_with_single_campaign(self) -> None:
        """Verify report works with single campaign."""
        snapshot = create_test_snapshot(1)
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            self.reporter.generate_report(snapshot, output_path)
            
            workbook = load_workbook(output_path)
            assert len(workbook.sheetnames) == 2

    def test_report_with_many_campaigns(self) -> None:
        """Verify report works with many campaigns."""
        snapshot = create_test_snapshot(20)
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            self.reporter.generate_report(snapshot, output_path)
            
            workbook = load_workbook(output_path)
            ws = workbook["Campaign Deep-Dive"]
            
            data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value)
            assert data_rows == 20

    def test_risk_status_values_present(self) -> None:
        """Verify risk status values appear in detail sheet."""
        snapshot = create_test_snapshot(3)
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            self.reporter.generate_report(snapshot, output_path)
            
            workbook = load_workbook(output_path)
            ws = workbook["Campaign Deep-Dive"]
            
            # Risk Status is column 9
            risk_values = [row[8].value for row in ws.iter_rows(min_row=2) if row[8].value]
            
            assert "HEALTHY" in risk_values
            assert "WARNING" in risk_values
            assert "CRITICAL" in risk_values
