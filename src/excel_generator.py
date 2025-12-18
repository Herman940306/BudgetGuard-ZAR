"""
BudgetGuard ZAR - Excel Report Generation Module.

This module generates professional Excel reports for budget analysis.
Follows the 'Executive First' principle with an at-a-glance Finance Summary
tab and a detailed Campaign Deep-Dive tab.

South African Market Context:
    - Native ZAR currency formatting (R #,##0.00)
    - Conditional formatting for risk levels
    - Professional styling for client presentations

Classes:
    ExcelReporter: Generates Excel workbooks from analysis snapshots.
"""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.schema import AnalysisSnapshot, RiskLevel


class ExcelReporter:
    """
    Generates professional Excel reports for budget analysis.

    Creates workbooks with Finance Summary and Campaign Deep-Dive
    sheets, applying ZAR currency formatting and risk-based
    conditional formatting.

    Attributes:
        ZAR_FORMAT: Excel number format for ZAR currency.
        PERCENTAGE_FORMAT: Excel number format for percentages.

    Example:
        >>> reporter = ExcelReporter()
        >>> reporter.generate_report(snapshot, "budget_report.xlsx")
    """

    # Excel number formats
    ZAR_FORMAT = 'R #,##0.00'
    PERCENTAGE_FORMAT = '0.00%'

    # Conditional formatting colours
    CRITICAL_FILL = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )
    WARNING_FILL = PatternFill(
        start_color="FFEB9C",
        end_color="FFEB9C",
        fill_type="solid"
    )
    HEALTHY_FILL = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    # Header styling
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(
        start_color="2F5496",
        end_color="2F5496",
        fill_type="solid"
    )
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    # Border styling
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def __init__(self):
        """Initialises the ExcelReporter."""
        self._setup_styles()

    def _setup_styles(self) -> None:
        """Creates named styles for consistent formatting."""
        self._currency_style = NamedStyle(name="zar_currency")
        self._currency_style.number_format = self.ZAR_FORMAT

    def generate_report(
        self,
        snapshot: AnalysisSnapshot,
        output_path: Union[str, Path]
    ) -> None:
        """
        Generates a complete Excel report from analysis results.

        Creates a workbook with two sheets:
        1. Finance Summary - Executive dashboard
        2. Campaign Deep-Dive - Detailed per-campaign data

        Args:
            snapshot: Complete analysis snapshot.
            output_path: Path for the output .xlsx file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()

        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        # Create sheets
        self._create_summary_sheet(workbook, snapshot)
        self._create_detail_sheet(workbook, snapshot)

        workbook.save(output_path)

    def _create_summary_sheet(
        self,
        workbook: Workbook,
        snapshot: AnalysisSnapshot
    ) -> None:
        """
        Creates the Finance Summary sheet with aggregate metrics.

        Executive-first design with key metrics prominently displayed.

        Args:
            workbook: Target workbook.
            snapshot: Analysis data.
        """
        ws = workbook.create_sheet("Finance Summary")

        # Title
        ws["A1"] = "BudgetGuard ZAR - Finance Summary"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        # Report metadata
        ws["A3"] = "Report Generated:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws["A4"] = "Analysis Date:"
        ws["B4"] = snapshot.timestamp.strftime("%Y-%m-%d %H:%M")
        ws["A5"] = "Version:"
        ws["B5"] = snapshot.version

        # Portfolio Overview section
        ws["A7"] = "PORTFOLIO OVERVIEW"
        ws["A7"].font = Font(bold=True, size=14)
        ws.merge_cells("A7:D7")

        # Key metrics table
        metrics = [
            ("Total Portfolio Budget", snapshot.total_budget),
            ("Total Spend to Date", snapshot.total_spend),
            ("Remaining Budget", snapshot.total_budget - snapshot.total_spend),
            ("Overall RDS", self._calculate_overall_rds(snapshot)),
        ]

        row = 9
        for label, value in metrics:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = float(value)
            ws[f"B{row}"].number_format = self.ZAR_FORMAT
            row += 1

        # Risk Summary section
        ws["A14"] = "RISK SUMMARY"
        ws["A14"].font = Font(bold=True, size=14)
        ws.merge_cells("A14:D14")

        # Risk counts table
        healthy_count = sum(
            1 for ca in snapshot.campaigns
            if ca.risk_level == RiskLevel.HEALTHY
        )
        over_budget_count = sum(
            1 for ca in snapshot.campaigns
            if ca.risk_level == RiskLevel.OVER_BUDGET
        )

        risk_data = [
            ("Risk Level", "Count", "Status"),
            ("CRITICAL", snapshot.critical_count, "Immediate Action Required"),
            ("WARNING", snapshot.warning_count, "Monitor Closely"),
            ("HEALTHY", healthy_count, "On Track"),
            ("OVER BUDGET", over_budget_count, "Budget Exceeded"),
        ]

        row = 16
        for i, (level, count, status) in enumerate(risk_data):
            ws[f"A{row}"] = level
            ws[f"B{row}"] = count
            ws[f"C{row}"] = status

            if i == 0:  # Header row
                for col in ["A", "B", "C"]:
                    ws[f"{col}{row}"].font = self.HEADER_FONT
                    ws[f"{col}{row}"].fill = self.HEADER_FILL
                    ws[f"{col}{row}"].alignment = self.HEADER_ALIGNMENT
            else:
                # Apply risk-based formatting
                if level == "CRITICAL":
                    ws[f"A{row}"].fill = self.CRITICAL_FILL
                elif level == "WARNING":
                    ws[f"A{row}"].fill = self.WARNING_FILL
                elif level == "HEALTHY":
                    ws[f"A{row}"].fill = self.HEALTHY_FILL

            for col in ["A", "B", "C"]:
                ws[f"{col}{row}"].border = self.THIN_BORDER

            row += 1

        # Campaign count
        ws["A22"] = "Total Campaigns Analysed:"
        ws["A22"].font = Font(bold=True)
        ws["B22"] = len(snapshot.campaigns)

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_detail_sheet(
        self,
        workbook: Workbook,
        snapshot: AnalysisSnapshot
    ) -> None:
        """
        Creates the Campaign Deep-Dive sheet with per-campaign data.

        Args:
            workbook: Target workbook.
            snapshot: Analysis data.
        """
        ws = workbook.create_sheet("Campaign Deep-Dive")

        # Headers
        headers = [
            "Campaign",
            "Monthly Budget",
            "Current Spend",
            "Remaining",
            "RDS",
            "Spend %",
            "Time %",
            "Variance",
            "Risk Status",
            "Days Left",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # Data rows
        for row_idx, analysis in enumerate(snapshot.campaigns, start=2):
            campaign = analysis.campaign
            remaining = campaign.monthly_budget - campaign.current_spend
            variance = analysis.spend_percentage - analysis.time_percentage

            row_data = [
                campaign.name,
                float(campaign.monthly_budget),
                float(campaign.current_spend),
                float(remaining),
                float(analysis.rds),
                float(analysis.spend_percentage) / 100,
                float(analysis.time_percentage) / 100,
                float(variance),
                analysis.risk_level.value,
                analysis.days_remaining,
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER

                # Apply number formats
                if col_idx in [2, 3, 4, 5]:  # Currency columns
                    cell.number_format = self.ZAR_FORMAT
                elif col_idx in [6, 7]:  # Percentage columns
                    cell.number_format = self.PERCENTAGE_FORMAT
                elif col_idx == 8:  # Variance
                    cell.number_format = '0.00'

            # Apply risk-based row formatting
            risk_fill = self._get_risk_fill(analysis.risk_level)
            if risk_fill:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = risk_fill

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _calculate_overall_rds(self, snapshot: AnalysisSnapshot) -> Decimal:
        """
        Calculates the overall RDS across all campaigns.

        Args:
            snapshot: Analysis snapshot.

        Returns:
            Sum of all campaign RDS values.
        """
        return sum(ca.rds for ca in snapshot.campaigns)

    def _get_risk_fill(self, risk_level: RiskLevel) -> PatternFill:
        """
        Returns the appropriate fill colour for a risk level.

        Args:
            risk_level: Campaign risk level.

        Returns:
            PatternFill for the risk level, or None for HEALTHY.
        """
        if risk_level == RiskLevel.CRITICAL:
            return self.CRITICAL_FILL
        elif risk_level == RiskLevel.WARNING:
            return self.WARNING_FILL
        elif risk_level == RiskLevel.OVER_BUDGET:
            return self.CRITICAL_FILL
        return None  # HEALTHY - no special fill

    def _auto_adjust_columns(self, worksheet: Worksheet) -> None:
        """
        Auto-adjusts column widths based on content.

        Args:
            worksheet: Target worksheet.
        """
        for col_idx in range(1, worksheet.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for row_idx in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except (TypeError, AttributeError):
                    pass

            # Add padding and set minimum width
            adjusted_width = max(max_length + 2, 10)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def generate_filename(self, prefix: str = "budget_report") -> str:
        """
        Generates a timestamped filename for reports.

        Args:
            prefix: Filename prefix. Defaults to "budget_report".

        Returns:
            Filename like "budget_report_2024-12-18_143052.xlsx".
        """
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        return f"{prefix}_{timestamp}.xlsx"
